"""
Daily dashboard report for Passport Photo Converter.

Reads traffic + feedback from Google Sheets, builds an Excel workbook with
KPI summary, daily metrics, charts, and raw feedback, then emails the file.

Usage:
    python report.py                 # generate + email
    python report.py --no-email      # generate only (no email)

Secrets are read from .streamlit/secrets.toml (same as the app) with
env-var overrides for CI/cron:
    GOOGLE_SERVICE_ACCOUNT_JSON, GOOGLE_SHEET_ID,
    SMTP_EMAIL, SMTP_PASSWORD, REPORT_RECIPIENT
"""
from __future__ import annotations

import argparse
import json
import os
import smtplib
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import gspread
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    CharacterProperties,
    Font as DrawingFont,
    Paragraph,
    ParagraphProperties,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SECRETS_PATH = Path(__file__).parent / ".streamlit" / "secrets.toml"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_FILE = OUTPUT_DIR / "passport_photo_dashboard.xlsx"

# -- Palette (colorblind-friendly, high-contrast) --
CLR_PRIMARY = "1E3A5F"
CLR_VISITS = "0E7490"
CLR_PHOTOS = "16A34A"
CLR_FEEDBACK = "D97706"
CLR_LIGHT_BG = "F8FAFC"
CLR_BORDER = "E2E8F0"
CLR_HEADER_BG = "1E3A5F"
CLR_HEADER_FG = "FFFFFF"

HEADER_FONT = Font(bold=True, size=11, color=CLR_HEADER_FG)
HEADER_FILL = PatternFill(start_color=CLR_HEADER_BG, end_color=CLR_HEADER_BG, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
KPI_LABEL_FONT = Font(bold=True, size=11, color="374151")
KPI_VALUE_FONT = Font(bold=True, size=16, color=CLR_PRIMARY)
INSIGHT_FONT = Font(italic=True, size=10, color="64748B")
THIN_BORDER = Border(
    bottom=Side(style="thin", color=CLR_BORDER),
)


# =========================================================================
# Secrets
# =========================================================================

def _load_toml_secrets() -> dict[str, str]:
    if not SECRETS_PATH.exists():
        return {}
    text = SECRETS_PATH.read_text()
    out: dict[str, str] = {}
    i = 0
    lines = text.splitlines()
    while i < len(lines):
        line = lines[i].strip()
        if "=" not in line or line.startswith("#"):
            i += 1
            continue
        key, _, val = line.partition("=")
        key = key.strip()
        val = val.strip()
        if val.startswith("'''") or val.startswith('"""'):
            delim = val[:3]
            parts = [val[3:]]
            i += 1
            while i < len(lines):
                if delim in lines[i]:
                    parts.append(lines[i].split(delim)[0])
                    break
                parts.append(lines[i])
                i += 1
            out[key] = "\n".join(parts)
        else:
            out[key] = val.strip("'").strip('"')
        i += 1
    return out


def _secret(key: str) -> str | None:
    val = os.environ.get(key)
    if val:
        return val
    return _load_toml_secrets().get(key)


def _service_account() -> tuple[dict, str]:
    raw = _secret("GOOGLE_SERVICE_ACCOUNT_JSON")
    sheet_id = _secret("GOOGLE_SHEET_ID")
    if not raw or not sheet_id:
        sys.exit("Missing GOOGLE_SERVICE_ACCOUNT_JSON or GOOGLE_SHEET_ID")
    if isinstance(raw, str):
        raw = raw.strip().strip("'").strip('"').strip()
        sa = json.loads(raw)
    else:
        sa = dict(raw)
    return sa, sheet_id


# =========================================================================
# Google Sheets data fetch
# =========================================================================

def _fetch_rows(sheet_id: str, client: gspread.Client, name: str) -> list[dict]:
    try:
        ws = client.open_by_key(sheet_id).worksheet(name)
        return ws.get_all_records()
    except gspread.WorksheetNotFound:
        return []


def _parse_date(iso_str: str) -> date | None:
    try:
        return datetime.fromisoformat(iso_str.replace("Z", "+00:00")).date()
    except Exception:
        return None


# =========================================================================
# KPI computation
# =========================================================================

def compute_metrics(
    traffic: list[dict], feedback: list[dict],
) -> tuple[dict, list[tuple], list[tuple]]:
    today = date.today()

    visits_by_day: Counter[date] = Counter()
    photos_by_day: Counter[date] = Counter()
    feedback_by_day: Counter[date] = Counter()
    unique_sessions: set[str] = set()
    today_sessions: set[str] = set()

    for row in traffic:
        d = _parse_date(str(row.get("submitted_at_utc", "")))
        event = row.get("event_name", "")
        sid = str(row.get("session_id", ""))
        if d is None:
            continue
        if event == "app_visit":
            visits_by_day[d] += 1
            unique_sessions.add(sid)
            if d == today:
                today_sessions.add(sid)
        elif event == "photo_processed":
            photos_by_day[d] += 1

    for row in feedback:
        d = _parse_date(str(row.get("submitted_at_utc", "")))
        if d:
            feedback_by_day[d] += 1

    total_visits = sum(visits_by_day.values())
    total_photos = sum(photos_by_day.values())
    total_feedback = len(feedback)
    conversion_rate = (total_photos / total_visits * 100) if total_visits else 0

    kpis = {
        "Report Date": today.isoformat(),
        "Total Visits (all time)": total_visits,
        "Total Photos Processed": total_photos,
        "Conversion Rate": f"{conversion_rate:.1f}%",
        "Total Feedback Entries": total_feedback,
        "Unique Sessions (all time)": len(unique_sessions),
        "Today's Active Users": len(today_sessions),
        "Today's Visits": visits_by_day.get(today, 0),
        "Today's Photos": photos_by_day.get(today, 0),
    }

    all_days = sorted(set(visits_by_day) | set(photos_by_day) | set(feedback_by_day))
    daily = [
        (d.isoformat(), visits_by_day[d], photos_by_day[d], feedback_by_day[d])
        for d in all_days
    ]

    fb_rows = []
    for row in feedback:
        d = _parse_date(str(row.get("submitted_at_utc", "")))
        fb_rows.append((
            d.isoformat() if d else str(row.get("submitted_at_utc", "")),
            str(row.get("feedback", "")),
        ))

    return kpis, daily, fb_rows


# =========================================================================
# Excel workbook
# =========================================================================

def _style_header(ws, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.row_dimensions[1].height = 28


def _chart_title(text: str) -> RichText:
    cp = CharacterProperties(
        latin=DrawingFont(typeface="Calibri"),
        sz=1200, b=True, solidFill=CLR_PRIMARY,
    )
    return RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])


def _style_chart(chart, title: str, color: str, is_line: bool = False) -> None:
    chart.width = 32
    chart.height = 16
    chart.title = title
    chart.style = 2
    chart.legend.position = "b"

    chart.y_axis.majorGridlines.spPr = GraphicalProperties(
        ln=LineProperties(solidFill="E5E7EB", w=6350)
    )
    chart.y_axis.tickLblPos = "low"
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    chart.y_axis.title = None
    chart.x_axis.title = None

    chart.plot_area.graphicalProperties = GraphicalProperties()
    chart.plot_area.graphicalProperties.noFill = True
    chart.plot_area.graphicalProperties.line = LineProperties(noFill=True)

    series = chart.series[0]
    if is_line:
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 28000
        series.smooth = True
    else:
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line = LineProperties(noFill=True)


def _add_data_labels(chart, show_val: bool = True) -> None:
    dl = DataLabelList()
    dl.showVal = show_val
    dl.showCatName = False
    dl.showSerName = False
    dl.numFmt = "#,##0"
    chart.series[0].dLbls = dl


def _compute_insights(daily: list[tuple]) -> dict[str, str]:
    if not daily:
        return {"visits": "", "photos": "", "feedback": ""}

    dates = [r[0] for r in daily]
    visits = [r[1] for r in daily]
    photos = [r[2] for r in daily]
    feedbacks = [r[3] for r in daily]

    peak_visit_idx = visits.index(max(visits))
    peak_photo_idx = photos.index(max(photos))
    total_fb = sum(feedbacks)

    v_insight = f"Peak: {max(visits)} visits on {dates[peak_visit_idx]}."
    if len(visits) >= 2:
        diff = visits[-1] - visits[-2]
        direction = "up" if diff > 0 else ("down" if diff < 0 else "flat")
        v_insight += f" Latest trend: {direction} ({diff:+d} vs prior day)."

    p_insight = f"Peak: {max(photos)} photos on {dates[peak_photo_idx]}. Total: {sum(photos)} processed."
    if sum(visits) > 0:
        rate = sum(photos) / sum(visits) * 100
        p_insight += f" Conversion rate: {rate:.1f}%."

    f_insight = f"{total_fb} feedback entries total."
    if total_fb > 0:
        peak_fb_idx = feedbacks.index(max(feedbacks))
        f_insight += f" Busiest day: {dates[peak_fb_idx]} ({max(feedbacks)})."

    return {"visits": v_insight, "photos": p_insight, "feedback": f_insight}


from openpyxl.chart.shapes import GraphicalProperties


def build_workbook(
    kpis: dict, daily: list[tuple], feedback_rows: list[tuple],
) -> Workbook:
    wb = Workbook()
    insights = _compute_insights(daily)

    # ── Sheet 1: KPI Summary ──
    ws = wb.active
    ws.title = "KPI Summary"
    ws.sheet_properties.tabColor = CLR_PRIMARY
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28

    title_cell = ws.cell(row=1, column=1, value="Passport Photo Converter — Daily Report")
    title_cell.font = Font(bold=True, size=18, color=CLR_PRIMARY)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 36

    ws.cell(row=2, column=1, value="Key Performance Indicators").font = Font(
        size=10, italic=True, color="94A3B8"
    )
    ws.merge_cells("A2:B2")

    row = 4
    for label, value in kpis.items():
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = KPI_LABEL_FONT
        lc.border = THIN_BORDER
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = KPI_VALUE_FONT
        vc.alignment = Alignment(horizontal="right")
        vc.border = THIN_BORDER
        row += 1

    # ── Sheet 2: Daily Metrics ──
    ws_daily = wb.create_sheet("Daily Metrics")
    ws_daily.sheet_properties.tabColor = CLR_PHOTOS
    headers = ["Date", "Visits", "Photos Processed", "Feedback"]
    ws_daily.append(headers)
    _style_header(ws_daily, len(headers))

    stripe = PatternFill(start_color=CLR_LIGHT_BG, end_color=CLR_LIGHT_BG, fill_type="solid")
    for i, r in enumerate(daily):
        ws_daily.append(list(r))
        if i % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws_daily.cell(row=i + 2, column=c).fill = stripe

    # ── Sheet 3: Charts ──
    ws_charts = wb.create_sheet("Charts")
    ws_charts.sheet_properties.tabColor = CLR_FEEDBACK

    if len(daily) >= 2:
        data_rows = len(daily) + 1
        cats = Reference(ws_daily, min_col=1, min_row=2, max_row=data_rows)
        chart_row = 1
        chart_spacing = 19

        # Chart 1: Daily Visits (line)
        c1 = LineChart()
        vals1 = Reference(ws_daily, min_col=2, min_row=1, max_row=data_rows)
        c1.add_data(vals1, titles_from_data=True)
        c1.set_categories(cats)
        _style_chart(c1, "Daily Visits", CLR_VISITS, is_line=True)
        ws_charts.add_chart(c1, f"A{chart_row}")

        insight_row = chart_row + chart_spacing
        ic = ws_charts.cell(row=insight_row, column=1, value=insights["visits"])
        ic.font = INSIGHT_FONT
        ws_charts.merge_cells(f"A{insight_row}:G{insight_row}")
        chart_row = insight_row + 2

        # Chart 2: Photos Processed (bar)
        c2 = BarChart()
        vals2 = Reference(ws_daily, min_col=3, min_row=1, max_row=data_rows)
        c2.add_data(vals2, titles_from_data=True)
        c2.set_categories(cats)
        _style_chart(c2, "Daily Photos Processed", CLR_PHOTOS)
        _add_data_labels(c2)
        ws_charts.add_chart(c2, f"A{chart_row}")

        insight_row = chart_row + chart_spacing
        ic2 = ws_charts.cell(row=insight_row, column=1, value=insights["photos"])
        ic2.font = INSIGHT_FONT
        ws_charts.merge_cells(f"A{insight_row}:G{insight_row}")
        chart_row = insight_row + 2

        # Chart 3: Feedback Volume (bar)
        c3 = BarChart()
        vals3 = Reference(ws_daily, min_col=4, min_row=1, max_row=data_rows)
        c3.add_data(vals3, titles_from_data=True)
        c3.set_categories(cats)
        _style_chart(c3, "Daily Feedback Volume", CLR_FEEDBACK)
        _add_data_labels(c3)
        ws_charts.add_chart(c3, f"A{chart_row}")

        insight_row = chart_row + chart_spacing
        ic3 = ws_charts.cell(row=insight_row, column=1, value=insights["feedback"])
        ic3.font = INSIGHT_FONT
        ws_charts.merge_cells(f"A{insight_row}:G{insight_row}")
    else:
        ws_charts.cell(row=1, column=1, value="Not enough data for charts (need at least 2 days).").font = INSIGHT_FONT

    # ── Sheet 4: Raw Feedback ──
    ws_fb = wb.create_sheet("Raw Feedback")
    ws_fb.sheet_properties.tabColor = "EF4444"
    fb_headers = ["Date", "Feedback"]
    ws_fb.append(fb_headers)
    _style_header(ws_fb, len(fb_headers))
    ws_fb.column_dimensions["B"].width = 80
    for i, r in enumerate(feedback_rows):
        ws_fb.append(list(r))
        if i % 2 == 0:
            for c in range(1, 3):
                ws_fb.cell(row=i + 2, column=c).fill = stripe

    return wb


# =========================================================================
# Email
# =========================================================================

def send_email(filepath: Path, recipient: str, smtp_email: str, smtp_password: str) -> None:
    msg = MIMEMultipart()
    msg["From"] = smtp_email
    msg["To"] = recipient
    msg["Subject"] = f"Passport Photo Dashboard — {date.today().isoformat()}"

    body = (
        f"Hi,\n\n"
        f"Attached is the daily dashboard report for {date.today().isoformat()}.\n\n"
        f"— Passport Photo Converter\n"
    )
    msg.attach(MIMEText(body, "plain"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={filepath.name}")
    msg.attach(part)

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(smtp_email, smtp_password)
        server.send_message(msg)

    print(f"Email sent to {recipient}")


# =========================================================================
# Main
# =========================================================================

def main() -> None:
    parser = argparse.ArgumentParser(description="Generate & email the daily dashboard report.")
    parser.add_argument("--no-email", action="store_true", help="Skip sending the email")
    args = parser.parse_args()

    print("Connecting to Google Sheets...")
    sa, sheet_id = _service_account()
    client = gspread.service_account_from_dict(sa)

    traffic_ws = _secret("GOOGLE_TRAFFIC_WORKSHEET") or "traffic"
    feedback_ws = _secret("GOOGLE_SHEET_WORKSHEET") or "feedback"

    print(f"Fetching traffic from '{traffic_ws}'...")
    traffic = _fetch_rows(sheet_id, client, traffic_ws)
    print(f"  → {len(traffic)} rows")

    print(f"Fetching feedback from '{feedback_ws}'...")
    feedback = _fetch_rows(sheet_id, client, feedback_ws)
    print(f"  → {len(feedback)} rows")

    print("Computing KPIs...")
    kpis, daily, fb_rows = compute_metrics(traffic, feedback)
    for k, v in kpis.items():
        print(f"  {k}: {v}")

    print("Building workbook...")
    wb = build_workbook(kpis, daily, fb_rows)
    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(str(OUTPUT_FILE))
    print(f"Saved → {OUTPUT_FILE}")

    if args.no_email:
        print("Skipping email (--no-email).")
        return

    smtp_email = _secret("SMTP_EMAIL")
    smtp_password = _secret("SMTP_PASSWORD")
    recipient = _secret("REPORT_RECIPIENT") or "supportpassportphotoconversion@gmail.com"

    if not smtp_email or not smtp_password:
        print("SMTP_EMAIL / SMTP_PASSWORD not set — skipping email.")
        print("Add them to .streamlit/secrets.toml or set as env vars.")
        return

    print(f"Emailing to {recipient}...")
    send_email(OUTPUT_FILE, recipient, smtp_email, smtp_password)
    print("Done.")


if __name__ == "__main__":
    main()
